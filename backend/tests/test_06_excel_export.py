"""Excel 출력 API 테스트"""
import io
import pytest
import httpx
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from tests.conftest import auth_header


def test_export_approved_dn(client: httpx.Client, admin_token: str):
    """APPROVED/EXPORTED DN Excel 출력 + 구조 검증"""
    list_res = client.get("/api/v1/debit-notes", headers=auth_header(admin_token))
    dns = [d for d in list_res.json()["items"] if d["status"] in ("APPROVED", "EXPORTED")]
    if not dns:
        pytest.skip("No APPROVED/EXPORTED DN")

    dn_id = dns[0]["debit_note_id"]
    res = client.post(f"/api/v1/debit-notes/{dn_id}/export-excel", headers=auth_header(admin_token))
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers.get("content-type", "")

    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A1"].value == "UNI CONSULTING CO.LTD"
    assert ws["C5"].value == "TO"
    assert ws["C9"].value == "Exchange rate"
    assert ws["A12"].value == "DEBIT NOTE"


def test_export_draft_dn_rejected(client: httpx.Client, admin_token: str):
    """DRAFT DN은 출력 거부 (400)"""
    list_res = client.get("/api/v1/debit-notes?status=DRAFT", headers=auth_header(admin_token))
    drafts = list_res.json()["items"]
    if not drafts:
        pytest.skip("No DRAFT DN")
    dn_id = drafts[0]["debit_note_id"]
    res = client.post(f"/api/v1/debit-notes/{dn_id}/export-excel", headers=auth_header(admin_token))
    assert res.status_code == 400


def test_excel_formulas_import(client: httpx.Client, admin_token: str):
    """IMPORT 시트 수식 검증"""
    list_res = client.get("/api/v1/debit-notes", headers=auth_header(admin_token))
    exported = [d for d in list_res.json()["items"] if d["status"] in ("APPROVED", "EXPORTED")]
    if not exported:
        pytest.skip("No exportable DN")

    dn_id = exported[0]["debit_note_id"]
    res = client.post(f"/api/v1/debit-notes/{dn_id}/export-excel", headers=auth_header(admin_token))
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active

    bc = str(ws.cell(row=16, column=column_index_from_string("BC")).value or "")
    bd = str(ws.cell(row=16, column=column_index_from_string("BD")).value or "")
    be = str(ws.cell(row=16, column=column_index_from_string("BE")).value or "")
    bf = str(ws.cell(row=16, column=column_index_from_string("BF")).value or "")

    assert "SUM(M16" in bc, f"BC16 수식 오류: {bc}"
    assert "$BE$13" in bd, f"BD16 수식 오류: {bd}"
    assert "8%" in be, f"BE16 수식 오류: {be}"
    assert "BD16" in bf or "BE16" in bf, f"BF16 수식 오류: {bf}"

    # 숨김 행 13
    assert ws.row_dimensions[13].hidden is True
    assert ws.cell(row=13, column=column_index_from_string("BE")).value is not None


def test_export_history(client: httpx.Client, admin_token: str):
    """출력 이력 조회"""
    list_res = client.get("/api/v1/debit-notes", headers=auth_header(admin_token))
    exported = [d for d in list_res.json()["items"] if d["status"] == "EXPORTED"]
    if not exported:
        pytest.skip("No EXPORTED DN")

    dn_id = exported[0]["debit_note_id"]
    res = client.get(f"/api/v1/debit-notes/{dn_id}/exports", headers=auth_header(admin_token))
    assert res.status_code == 200
    exports = res.json()
    assert len(exports) >= 1
    assert exports[0]["export_status"] == "COMPLETED"


def test_download_latest(client: httpx.Client, admin_token: str):
    """최근 Excel 다운로드"""
    list_res = client.get("/api/v1/debit-notes", headers=auth_header(admin_token))
    exported = [d for d in list_res.json()["items"] if d["status"] == "EXPORTED"]
    if not exported:
        pytest.skip("No EXPORTED DN")

    dn_id = exported[0]["debit_note_id"]
    res = client.get(f"/api/v1/debit-notes/{dn_id}/download", headers=auth_header(admin_token))
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers.get("content-type", "")
