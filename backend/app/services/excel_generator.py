"""Excel Debit Note 생성 서비스 (설계서 FR-021 ~ FR-028)

NEXCON 기술사양서 기반:
- IMPORT Sheet: A-BM (65 columns), Row 16+
- EXPORT Sheet: A-AQ (43 columns), Row 16+
- 수식: BC=SUM(M:AT), BD=BC*환율, BE=SUM(Z:AT)*환율*8%, BF=BD+BE
"""
import os
import io
from datetime import date, datetime
from decimal import Decimal
from collections import defaultdict
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers,
)
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.debit_note import DebitNote, DebitNoteLine
from app.models.shipment import Shipment, ShipmentFeeDetail
from app.models.fee import FeeItem
from app.models.client import Client, ClientTemplate, ClientFeeMapping


# ── 스타일 상수 ──────────────────────────────────────────────
TITLE_FONT = Font(name="Arial", size=14, bold=True)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=9)
COMPANY_FONT = Font(name="Arial", size=11, bold=True)
LABEL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DUPLICATE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NUMBER_FMT_USD = '#,##0.00'
NUMBER_FMT_VND = '#,##0'

# ── IMPORT 컬럼 헤더 정의 ────────────────────────────────────
IMPORT_HEADERS = {
    "A": "No.", "B": "Delivery Date", "C": "Commercial Invoice No.",
    "D": "MBL", "E": "HBL", "F": "Term",
    "G": "No. of pkgs", "H": "Gross weight", "I": "Chargable weight",
    "J": "CD No.", "K": "CD Type", "L": "Air rate/Ocean freight",
    # M-AT: 비용 항목 (동적)
    "BA": "Pay on behalf",
    "BC": "Total (USD)", "BD": "Total (VND)",
    "BE": "VAT (8%)", "BF": "Grand total (VND)",
    "BG": "Back-to-back Invoice",
}

EXPORT_HEADERS = {
    "A": "No.", "B": "Delivery Date", "C": "Commercial Invoice No.",
    "D": "MBL", "E": "HBL", "F": "Term",
    "G": "No. of pkgs", "H": "Gross weight", "I": "Chargable weight",
    "M": "Origin/Destination",
    # N-AH: 비용 항목 (동적)
    "AI": "Subtotal", "AJ": "Total (VND)",
    "AK": "VAT (8%)", "AL": "Grand total (VND)",
}


async def generate_debit_note_excel(
    debit_note_id: int,
    db: AsyncSession,
) -> tuple[io.BytesIO, str]:
    """Debit Note를 Excel 파일로 생성

    Returns:
        (BytesIO buffer, filename)
    """
    # ── 1. 데이터 로드 ─────────────────────────────────────
    dn = (await db.execute(
        select(DebitNote)
        .options(
            selectinload(DebitNote.lines),
            selectinload(DebitNote.client),
        )
        .where(DebitNote.debit_note_id == debit_note_id)
    )).scalar_one_or_none()

    if not dn:
        raise ValueError(f"Debit Note {debit_note_id} not found")

    client = dn.client

    # 템플릿 로드
    templates = (await db.execute(
        select(ClientTemplate)
        .where(ClientTemplate.client_id == client.client_id, ClientTemplate.is_active == True)
    )).scalars().all()

    # fee 매핑 로드
    fee_mappings = (await db.execute(
        select(ClientFeeMapping)
        .options(selectinload(ClientFeeMapping.fee_item))
        .where(ClientFeeMapping.client_id == client.client_id, ClientFeeMapping.is_active == True)
        .order_by(ClientFeeMapping.sort_order)
    )).scalars().all()

    # 모든 fee_items 로드
    all_fee_items = (await db.execute(
        select(FeeItem).where(FeeItem.is_active == True).order_by(FeeItem.sort_order)
    )).scalars().all()
    fee_item_map = {fi.fee_item_id: fi for fi in all_fee_items}

    # 라인별 shipment + fee_details 로드
    shipment_ids = [line.shipment_id for line in dn.lines]
    shipments_result = await db.execute(
        select(Shipment)
        .options(selectinload(Shipment.fee_details).selectinload(ShipmentFeeDetail.fee_item))
        .where(Shipment.shipment_id.in_(shipment_ids))
    )
    shipments_by_id = {s.shipment_id: s for s in shipments_result.scalars().unique().all()}

    # 라인을 line_no 순서로 정렬
    lines_sorted = sorted(dn.lines, key=lambda l: l.line_no or 0)

    # ── 2. 워크북 생성 ─────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    exchange_rate = float(dn.exchange_rate or 26446)
    period_str = dn.period_from.strftime("%m%Y") if dn.period_from else date.today().strftime("%m%Y")

    # IMPORT/EXPORT 시트 분리
    import_lines = []
    export_lines = []
    for line in lines_sorted:
        ship = shipments_by_id.get(line.shipment_id)
        if ship:
            if ship.shipment_type == "EXPORT":
                export_lines.append((line, ship))
            else:
                import_lines.append((line, ship))

    sheets_to_create = []
    if dn.sheet_type in ("IMPORT", "ALL") and import_lines:
        imp_tmpl = next((t for t in templates if t.sheet_type == "IMPORT"), None)
        sheets_to_create.append(("IMPORT", import_lines, imp_tmpl))
    if dn.sheet_type in ("EXPORT", "ALL") and export_lines:
        exp_tmpl = next((t for t in templates if t.sheet_type == "EXPORT"), None)
        sheets_to_create.append(("EXPORT", export_lines, exp_tmpl))

    # 데이터가 없으면 기본 IMPORT 시트라도 생성
    if not sheets_to_create:
        imp_tmpl = next((t for t in templates if t.sheet_type == "IMPORT"), None)
        sheets_to_create.append(("IMPORT", import_lines or [], imp_tmpl))

    for sheet_type, lines_data, template in sheets_to_create:
        _build_sheet(
            wb=wb,
            sheet_type=sheet_type,
            lines_data=lines_data,
            template=template,
            client=client,
            dn=dn,
            exchange_rate=exchange_rate,
            period_str=period_str,
            fee_mappings=[fm for fm in fee_mappings if fm.sheet_type == sheet_type],
            fee_item_map=fee_item_map,
            all_fee_items=all_fee_items,
        )

    # ── 3. 출력 ──────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    dn_number = dn.debit_note_number or f"DN-{dn.debit_note_id}"
    filename = f"{client.client_code}_{dn_number}_{period_str}.xlsx"
    filename = filename.replace(" ", "_")

    return buffer, filename


def _build_sheet(
    wb: Workbook,
    sheet_type: str,
    lines_data: list,
    template: Optional[ClientTemplate],
    client: Client,
    dn: DebitNote,
    exchange_rate: float,
    period_str: str,
    fee_mappings: list,
    fee_item_map: dict,
    all_fee_items: list,
):
    """개별 시트 생성 (IMPORT 또는 EXPORT)"""

    # 시트명
    if template and template.sheet_name_pattern:
        sheet_name = template.sheet_name_pattern.replace("{MMYYYY}", period_str)
    else:
        sheet_name = f"{sheet_type} {client.client_code[:3]} {period_str}"
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel 시트명 31자 제한

    data_start_row = template.data_start_row if template else 16
    is_import = sheet_type == "IMPORT"

    # ── 비용 컬럼 구성 ───────────────────────────────────
    # fee_mappings가 있으면 사용, 없으면 데이터에서 사용된 fee_item 기반으로 자동 구성
    fee_col_map = {}  # fee_item_id -> column_letter
    col_fee_map = {}  # column_letter -> fee_item

    if fee_mappings:
        for fm in fee_mappings:
            fee_col_map[fm.fee_item_id] = fm.column_letter
            col_fee_map[fm.column_letter] = fm.fee_item
    else:
        # 데이터에서 사용된 fee_item들을 자동으로 컬럼 배치
        used_fee_ids = set()
        for line, ship in lines_data:
            for fd in ship.fee_details:
                used_fee_ids.add(fd.fee_item_id)

        start_col = "M"
        start_idx = column_index_from_string(start_col)
        # VAT 0% (freight) 먼저, 그 다음 VAT 8% (local)
        non_vat_items = [fi for fi in all_fee_items if fi.fee_item_id in used_fee_ids and not fi.is_vat_applicable]
        vat_items = [fi for fi in all_fee_items if fi.fee_item_id in used_fee_ids and fi.is_vat_applicable]

        col_idx = start_idx
        for fi in non_vat_items + vat_items:
            col_letter = get_column_letter(col_idx)
            fee_col_map[fi.fee_item_id] = col_letter
            col_fee_map[col_letter] = fi
            col_idx += 1

    # 합계 컬럼 위치 결정
    if is_import:
        total_usd_col = template.total_usd_column if template else "BC"
        total_vnd_col = template.total_vnd_column if template else "BD"
        vat_col = template.vat_column if template else "BE"
        grand_total_col = template.grand_total_column if template else "BF"
        fee_start = template.fee_column_start if template else "M"
        fee_end = template.fee_column_end if template else "AT"
    else:
        total_usd_col = template.total_usd_column if template else "AI"
        total_vnd_col = template.total_vnd_column if template else "AJ"
        vat_col = template.vat_column if template else "AK"
        grand_total_col = template.grand_total_column if template else "AL"
        fee_start = template.fee_column_start if template else "M"
        fee_end = template.fee_column_end if template else "AH"

    # ── 헤더 영역 (Row 1-15) ─────────────────────────────
    _build_header(
        ws=ws,
        client=client,
        dn=dn,
        exchange_rate=exchange_rate,
        is_import=is_import,
        total_usd_col=total_usd_col,
        grand_total_col=grand_total_col,
        vat_col=vat_col,
    )

    # ── 컬럼 헤더 (Row 14-15) ────────────────────────────
    base_headers = IMPORT_HEADERS if is_import else EXPORT_HEADERS
    # 기본 헤더 쓰기
    for col_letter, header_text in base_headers.items():
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=14, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        # 14-15행 병합
        ws.merge_cells(start_row=14, start_column=col_idx, end_row=15, end_column=col_idx)

    # 비용 컬럼 헤더
    for col_letter, fi in col_fee_map.items():
        col_idx = column_index_from_string(col_letter)
        display_name = fi.item_name if hasattr(fi, 'item_name') else str(fi)
        cell = ws.cell(row=14, column=col_idx, value=display_name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.merge_cells(start_row=14, start_column=col_idx, end_row=15, end_column=col_idx)

    # ── 데이터 행 (Row 16+) ──────────────────────────────
    # 중복 감지용 딕셔너리
    hbl_counts = defaultdict(int)
    mbl_counts = defaultdict(int)
    inv_counts = defaultdict(int)
    cd_counts = defaultdict(int)

    for line, ship in lines_data:
        if ship.hbl:
            hbl_counts[ship.hbl] += 1
        if ship.mbl:
            mbl_counts[ship.mbl] += 1
        if ship.invoice_no:
            inv_counts[ship.invoice_no] += 1
        if ship.cd_no:
            cd_counts[ship.cd_no] += 1

    # 사용된 컬럼 추적 (빈 컬럼 숨김용)
    used_columns = set()

    sum_total_usd = Decimal("0")
    sum_total_vnd = 0
    sum_vat = 0
    sum_grand_total = 0

    for idx, (line, ship) in enumerate(lines_data):
        row = data_start_row + idx
        fee_details_map = {fd.fee_item_id: fd for fd in ship.fee_details}

        # A: No.
        ws.cell(row=row, column=1, value=idx + 1).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        used_columns.add("A")

        # B: Delivery Date
        if ship.delivery_date:
            cell = ws.cell(row=row, column=2, value=ship.delivery_date)
            cell.number_format = "YYYY-MM-DD"
            cell.font = DATA_FONT
            used_columns.add("B")

        # C: Invoice No
        if ship.invoice_no:
            cell = ws.cell(row=row, column=3, value=ship.invoice_no)
            cell.font = DATA_FONT
            if inv_counts.get(ship.invoice_no, 0) > 1:
                cell.fill = DUPLICATE_FILL
            used_columns.add("C")

        # D: MBL
        if ship.mbl:
            cell = ws.cell(row=row, column=4, value=ship.mbl)
            cell.font = DATA_FONT
            if mbl_counts.get(ship.mbl, 0) > 1:
                cell.fill = DUPLICATE_FILL
            used_columns.add("D")

        # E: HBL
        if ship.hbl:
            cell = ws.cell(row=row, column=5, value=ship.hbl)
            cell.font = DATA_FONT
            if hbl_counts.get(ship.hbl, 0) > 1:
                cell.fill = DUPLICATE_FILL
            used_columns.add("E")

        # F: Term
        if ship.term:
            ws.cell(row=row, column=6, value=ship.term).font = DATA_FONT
            used_columns.add("F")

        # G: No. of pkgs
        if ship.no_of_pkgs:
            ws.cell(row=row, column=7, value=ship.no_of_pkgs).font = DATA_FONT
            used_columns.add("G")

        # H: Gross weight
        if ship.gross_weight:
            cell = ws.cell(row=row, column=8, value=float(ship.gross_weight))
            cell.font = DATA_FONT
            cell.number_format = NUMBER_FMT_USD
            used_columns.add("H")

        # I: Chargeable weight
        if ship.chargeable_weight:
            cell = ws.cell(row=row, column=9, value=float(ship.chargeable_weight))
            cell.font = DATA_FONT
            cell.number_format = NUMBER_FMT_USD
            used_columns.add("I")

        if is_import:
            # J: CD No.
            if ship.cd_no:
                cell = ws.cell(row=row, column=10, value=ship.cd_no)
                cell.font = DATA_FONT
                if cd_counts.get(ship.cd_no, 0) > 1:
                    cell.fill = DUPLICATE_FILL
                used_columns.add("J")

            # K: CD Type
            if ship.cd_type:
                ws.cell(row=row, column=11, value=ship.cd_type).font = DATA_FONT
                used_columns.add("K")

            # L: Air rate/Ocean freight
            if ship.air_ocean_rate:
                ws.cell(row=row, column=12, value=ship.air_ocean_rate).font = DATA_FONT
                used_columns.add("L")
        else:
            # EXPORT: M = Origin/Destination (handled via fee_col or direct)
            if ship.origin_destination:
                col_idx = column_index_from_string("M")
                ws.cell(row=row, column=col_idx, value=ship.origin_destination).font = DATA_FONT
                used_columns.add("M")

        # 비용 항목 데이터 (M-AT or M-AH)
        for fee_item_id, col_letter in fee_col_map.items():
            fd = fee_details_map.get(fee_item_id)
            if fd and fd.amount_usd:
                col_idx = column_index_from_string(col_letter)
                amount = float(fd.amount_usd)
                cell = ws.cell(row=row, column=col_idx, value=amount)
                cell.font = DATA_FONT
                cell.number_format = NUMBER_FMT_USD
                cell.alignment = Alignment(horizontal="right")
                used_columns.add(col_letter)

        # Pay on behalf (IMPORT: BA)
        if is_import and line.pay_on_behalf and float(line.pay_on_behalf) > 0:
            ba_idx = column_index_from_string("BA")
            cell = ws.cell(row=row, column=ba_idx, value=float(line.pay_on_behalf))
            cell.font = DATA_FONT
            cell.number_format = NUMBER_FMT_USD
            cell.alignment = Alignment(horizontal="right")
            used_columns.add("BA")

        # ── 수식 컬럼 (BC-BF / AI-AL) ────────────────────
        if is_import:
            # BC = SUM(M:AT)
            bc_idx = column_index_from_string(total_usd_col)
            fee_start_letter = fee_start
            fee_end_letter = fee_end
            ws.cell(row=row, column=bc_idx).value = f"=SUM({fee_start_letter}{row}:{fee_end_letter}{row})"
            ws.cell(row=row, column=bc_idx).number_format = NUMBER_FMT_USD
            ws.cell(row=row, column=bc_idx).font = DATA_FONT
            ws.cell(row=row, column=bc_idx).alignment = Alignment(horizontal="right")
            used_columns.add(total_usd_col)

            # BD = BC * 환율 ($BE$13)
            bd_idx = column_index_from_string(total_vnd_col)
            ws.cell(row=row, column=bd_idx).value = f"={total_usd_col}{row}*$BE$13"
            ws.cell(row=row, column=bd_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=bd_idx).font = DATA_FONT
            ws.cell(row=row, column=bd_idx).alignment = Alignment(horizontal="right")
            used_columns.add(total_vnd_col)

            # BE = SUM(Z:AT) * 환율 * 8%
            be_idx = column_index_from_string(vat_col)
            # VAT 적용 범위: Z-AT (현지비용만)
            ws.cell(row=row, column=be_idx).value = f"=SUM(Z{row}:{fee_end_letter}{row})*$BE$13*8%"
            ws.cell(row=row, column=be_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=be_idx).font = DATA_FONT
            ws.cell(row=row, column=be_idx).alignment = Alignment(horizontal="right")
            used_columns.add(vat_col)

            # BF = BD + BE
            bf_idx = column_index_from_string(grand_total_col)
            ws.cell(row=row, column=bf_idx).value = f"={total_vnd_col}{row}+{vat_col}{row}"
            ws.cell(row=row, column=bf_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=bf_idx).font = DATA_FONT
            ws.cell(row=row, column=bf_idx).alignment = Alignment(horizontal="right")
            used_columns.add(grand_total_col)

        else:
            # EXPORT 수식
            # AI = SUM(M:N) -- subtotal
            ai_idx = column_index_from_string(total_usd_col)
            ws.cell(row=row, column=ai_idx).value = f"=SUM(M{row}+N{row})"
            ws.cell(row=row, column=ai_idx).number_format = NUMBER_FMT_USD
            ws.cell(row=row, column=ai_idx).font = DATA_FONT
            ws.cell(row=row, column=ai_idx).alignment = Alignment(horizontal="right")
            used_columns.add(total_usd_col)

            # AJ = ROUND(SUM(M:AH) * 환율, 0)
            aj_idx = column_index_from_string(total_vnd_col)
            ws.cell(row=row, column=aj_idx).value = f"=ROUND(SUM(M{row}:{fee_end}{row})*$D$9,0)"
            ws.cell(row=row, column=aj_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=aj_idx).font = DATA_FONT
            ws.cell(row=row, column=aj_idx).alignment = Alignment(horizontal="right")
            used_columns.add(total_vnd_col)

            # AK = AJ * 8%
            ak_idx = column_index_from_string(vat_col)
            ws.cell(row=row, column=ak_idx).value = f"={total_vnd_col}{row}*8%"
            ws.cell(row=row, column=ak_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=ak_idx).font = DATA_FONT
            ws.cell(row=row, column=ak_idx).alignment = Alignment(horizontal="right")
            used_columns.add(vat_col)

            # AL = AJ + AK
            al_idx = column_index_from_string(grand_total_col)
            ws.cell(row=row, column=al_idx).value = f"=SUM({total_vnd_col}{row}+{vat_col}{row})"
            ws.cell(row=row, column=al_idx).number_format = NUMBER_FMT_VND
            ws.cell(row=row, column=al_idx).font = DATA_FONT
            ws.cell(row=row, column=al_idx).alignment = Alignment(horizontal="right")
            used_columns.add(grand_total_col)

        # BG: Back-to-back Invoice (IMPORT only)
        if is_import and ship.back_to_back_invoice:
            bg_idx = column_index_from_string("BG")
            ws.cell(row=row, column=bg_idx, value=ship.back_to_back_invoice).font = DATA_FONT
            used_columns.add("BG")

        # 테두리 적용 (데이터 행)
        _apply_row_borders(ws, row, is_import, used_columns)

        # 합계 누적
        sum_total_usd += line.total_usd or Decimal("0")
        sum_total_vnd += int(line.total_vnd or 0)
        sum_vat += int(line.vat_amount or 0)
        sum_grand_total += int(line.grand_total_vnd or 0)

    # ── 합계 행 ──────────────────────────────────────────
    if lines_data:
        sum_row = data_start_row + len(lines_data)
        last_data_row = data_start_row + len(lines_data) - 1

        # "TOTAL" 라벨
        ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center")

        # 합계 수식
        tc_idx = column_index_from_string(total_usd_col)
        ws.cell(row=sum_row, column=tc_idx).value = f"=SUM({total_usd_col}{data_start_row}:{total_usd_col}{last_data_row})"
        ws.cell(row=sum_row, column=tc_idx).number_format = NUMBER_FMT_USD
        ws.cell(row=sum_row, column=tc_idx).font = Font(name="Arial", size=10, bold=True)

        tv_idx = column_index_from_string(total_vnd_col)
        ws.cell(row=sum_row, column=tv_idx).value = f"=SUM({total_vnd_col}{data_start_row}:{total_vnd_col}{last_data_row})"
        ws.cell(row=sum_row, column=tv_idx).number_format = NUMBER_FMT_VND
        ws.cell(row=sum_row, column=tv_idx).font = Font(name="Arial", size=10, bold=True)

        vc_idx = column_index_from_string(vat_col)
        ws.cell(row=sum_row, column=vc_idx).value = f"=SUM({vat_col}{data_start_row}:{vat_col}{last_data_row})"
        ws.cell(row=sum_row, column=vc_idx).number_format = NUMBER_FMT_VND
        ws.cell(row=sum_row, column=vc_idx).font = Font(name="Arial", size=10, bold=True)

        gt_idx = column_index_from_string(grand_total_col)
        ws.cell(row=sum_row, column=gt_idx).value = f"=SUM({grand_total_col}{data_start_row}:{grand_total_col}{last_data_row})"
        ws.cell(row=sum_row, column=gt_idx).number_format = NUMBER_FMT_VND
        ws.cell(row=sum_row, column=gt_idx).font = Font(name="Arial", size=10, bold=True)

        # 비용 항목별 합계
        for col_letter in fee_col_map.values():
            if col_letter in used_columns:
                c_idx = column_index_from_string(col_letter)
                ws.cell(row=sum_row, column=c_idx).value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
                ws.cell(row=sum_row, column=c_idx).number_format = NUMBER_FMT_USD
                ws.cell(row=sum_row, column=c_idx).font = Font(name="Arial", size=10, bold=True)

        # 합계 행 테두리
        _apply_row_borders(ws, sum_row, is_import, used_columns)

    # ── 빈 컬럼 숨김 (FR-022) ────────────────────────────
    _hide_empty_columns(ws, is_import, used_columns, fee_col_map, total_usd_col, grand_total_col)

    # ── 컬럼 너비 조정 ──────────────────────────────────
    _adjust_column_widths(ws, is_import)

    # ── 행 고정 (Row 1-15) ───────────────────────────────
    ws.freeze_panes = f"A{data_start_row}"


def _build_header(
    ws, client, dn, exchange_rate, is_import,
    total_usd_col, grand_total_col, vat_col,
):
    """헤더 영역 (Row 1-13) 생성"""
    # Row 1-2: 회사명
    ws.cell(row=1, column=1, value="UNI CONSULTING CO.LTD").font = COMPANY_FONT
    ws.cell(row=2, column=1, value="Tax code: 0315609***").font = Font(name="Arial", size=9)

    # Row 5-8: 수신자 정보
    ws.cell(row=5, column=3, value="TO").font = LABEL_FONT
    ws.cell(row=5, column=4, value=client.client_name or client.client_code).font = LABEL_FONT

    ws.cell(row=6, column=3, value="ADD").font = LABEL_FONT
    ws.cell(row=6, column=4, value=client.address or "").font = LABEL_FONT

    ws.cell(row=7, column=3, value="SUBJECT").font = LABEL_FONT
    period_label = f"Debit Note {dn.period_from.strftime('%m/%Y')}" if dn.period_from else "Debit Note"
    ws.cell(row=7, column=4, value=period_label).font = LABEL_FONT

    ws.cell(row=8, column=3, value="DATE").font = LABEL_FONT
    ws.cell(row=8, column=4, value=dn.billing_date or date.today()).font = LABEL_FONT
    ws.cell(row=8, column=4).number_format = "YYYY-MM-DD"

    # Row 9: Exchange rate (D9)
    ws.cell(row=9, column=3, value="Exchange rate").font = LABEL_FONT
    ws.cell(row=9, column=4, value=exchange_rate).font = Font(name="Arial", size=10, bold=True, color="FF0000")
    ws.cell(row=9, column=4).number_format = NUMBER_FMT_VND

    # Row 12: DEBIT NOTE 타이틀 (FR-026)
    last_col = grand_total_col
    last_col_idx = column_index_from_string(last_col)
    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=last_col_idx)
    title_cell = ws.cell(row=12, column=1, value="DEBIT NOTE")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 13: 숨김 행 -- 환율 참조 (IMPORT 전용)
    if is_import:
        bd_col_idx = column_index_from_string("BD")
        ws.cell(row=13, column=bd_col_idx, value="Tỷ giá").font = Font(name="Arial", size=8, color="999999")
        be_col_idx = column_index_from_string("BE")
        ws.cell(row=13, column=be_col_idx, value=exchange_rate).font = Font(name="Arial", size=8, color="999999")
        ws.cell(row=13, column=be_col_idx).number_format = NUMBER_FMT_VND
        ws.row_dimensions[13].hidden = True


def _apply_row_borders(ws, row, is_import, used_columns):
    """행에 테두리 적용 (FR-024)"""
    for col_letter in used_columns:
        try:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        except (ValueError, KeyError):
            pass


def _hide_empty_columns(ws, is_import, used_columns, fee_col_map, total_usd_col, grand_total_col):
    """빈 컬럼 숨김 (FR-022)"""
    # 비용 컬럼 중 데이터가 없는 것 숨김
    fee_start_idx = column_index_from_string("M")
    fee_end_idx = column_index_from_string("AT" if is_import else "AH")

    summary_cols = {total_usd_col, grand_total_col}
    # 합계 컬럼은 숨기지 않음
    for col_idx in range(fee_start_idx, fee_end_idx + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter not in used_columns and col_letter not in summary_cols:
            ws.column_dimensions[col_letter].hidden = True


def _adjust_column_widths(ws, is_import):
    """컬럼 너비 조정"""
    width_map = {
        "A": 5, "B": 12, "C": 18, "D": 16, "E": 16,
        "F": 7, "G": 8, "H": 10, "I": 10,
    }
    if is_import:
        width_map.update({"J": 14, "K": 8, "L": 10})

    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width

    # 비용 컬럼 기본 너비
    fee_start = column_index_from_string("M")
    fee_end = column_index_from_string("BG" if is_import else "AQ")
    for ci in range(fee_start, fee_end + 1):
        cl = get_column_letter(ci)
        if cl not in width_map:
            ws.column_dimensions[cl].width = 12
